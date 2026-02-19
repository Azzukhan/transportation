from io import BytesIO

import pytest
from httpx import AsyncClient
from openpyxl import Workbook


async def _auth_headers(client: AsyncClient) -> dict[str, str]:
    token_response = await client.post(
        "/api/v1/auth/token",
        json={"username": "admin", "password": "secret"},
    )
    token = token_response.cookies.get("access_token")
    assert token
    return {"Authorization": f"Bearer {token}"}


def _build_salary_workbook(rows: list[dict[str, object]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for row_index, row in enumerate(rows, start=9):
        ws.cell(row=row_index, column=2, value=row.get("employee_name"))  # B
        ws.cell(row=row_index, column=3, value=row.get("work_permit_no"))  # C
        ws.cell(row=row_index, column=4, value=row.get("personal_no"))  # D
        ws.cell(row=row_index, column=5, value=row.get("bank_name_routing_no"))  # E
        ws.cell(row=row_index, column=6, value=row.get("bank_account_no"))  # F
        ws.cell(row=row_index, column=7, value=row.get("days_absent"))  # G
        ws.cell(row=row_index, column=8, value=row.get("fixed_portion"))  # H
        ws.cell(row=row_index, column=9, value=row.get("variable_portion"))  # I
        ws.cell(row=row_index, column=10, value=row.get("total_payment"))  # J
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


@pytest.mark.asyncio
async def test_import_employee_salaries_creates_and_updates(client: AsyncClient) -> None:
    headers = await _auth_headers(client)

    first_file = _build_salary_workbook(
        [
            {
                "employee_name": "Ali",
                "work_permit_no": "12345678",
                "personal_no": "12345678901234",
                "bank_name_routing_no": "Bank X / 001",
                "bank_account_no": "ACC001",
                "days_absent": 1,
                "fixed_portion": "1000.00",
                "variable_portion": "50.00",
                "total_payment": "1050.00",
            },
            {
                "employee_name": "Invalid Row",
                "work_permit_no": "12345678",
                "personal_no": None,
                "bank_name_routing_no": "Bank X / 001",
                "bank_account_no": "ACC009",
                "days_absent": 2,
                "fixed_portion": "200.00",
                "variable_portion": "20.00",
                "total_payment": "220.00",
            },
        ]
    )

    import_response = await client.post(
        "/api/v1/employee-salaries/import",
        headers=headers,
        files={
            "file": (
                "salary_import.xlsx",
                first_file,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert import_response.status_code == 200
    assert import_response.json()["created"] == 1
    assert import_response.json()["updated"] == 0
    assert "personal_no is required" in import_response.json()["skipped"][0]["reason"]

    second_file = _build_salary_workbook(
        [
            {
                "employee_name": "Ali Updated",
                "work_permit_no": "12345678",
                "personal_no": "12345678901234",
                "bank_name_routing_no": "Bank X / 001",
                "bank_account_no": "ACC001",
                "days_absent": 0,
                "fixed_portion": "1200.00",
                "variable_portion": "25.00",
                "total_payment": "1225.00",
            }
        ]
    )
    import_again = await client.post(
        "/api/v1/employee-salaries/import",
        headers=headers,
        files={
            "file": (
                "salary_import_2.xlsx",
                second_file,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert import_again.status_code == 200
    assert import_again.json()["created"] == 0
    assert import_again.json()["updated"] == 1

    rows = await client.get("/api/v1/employee-salaries", headers=headers)
    assert rows.status_code == 200
    payload = rows.json()
    assert len(payload) == 1
    assert payload[0]["employee_name"] == "Ali Updated"
    assert payload[0]["fixed_portion"] == "1200.00"
    assert payload[0]["total_payment"] == "1225.00"


@pytest.mark.asyncio
async def test_import_employee_salaries_rejects_non_excel_file(client: AsyncClient) -> None:
    headers = await _auth_headers(client)
    response = await client.post(
        "/api/v1/employee-salaries/import",
        headers=headers,
        files={"file": ("salary.txt", b"invalid", "text/plain")},
    )
    assert response.status_code == 400
    assert (
        response.json()["error"]["message"]
        == "Unsupported file format. Please upload an .xlsx file"
    )


@pytest.mark.asyncio
async def test_import_employee_salaries_left_pads_numeric_ids(client: AsyncClient) -> None:
    headers = await _auth_headers(client)
    file_bytes = _build_salary_workbook(
        [
            {
                "employee_name": "Padded IDs",
                "work_permit_no": 1234567,
                "personal_no": 1234567890123,
                "bank_name_routing_no": "Bank Y",
                "bank_account_no": "ACC777",
                "days_absent": 0,
                "fixed_portion": "500.00",
                "variable_portion": "50.00",
                "total_payment": "550.00",
            }
        ]
    )
    import_response = await client.post(
        "/api/v1/employee-salaries/import",
        headers=headers,
        files={
            "file": (
                "salary_import_pad.xlsx",
                file_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert import_response.status_code == 200
    assert import_response.json()["created"] == 1

    rows = await client.get("/api/v1/employee-salaries", headers=headers)
    assert rows.status_code == 200
    payload = rows.json()
    assert len(payload) == 1
    assert payload[0]["work_permit_no"] == "01234567"
    assert payload[0]["personal_no"] == "01234567890123"


@pytest.mark.asyncio
async def test_import_employee_salaries_accepts_scientific_notation_ids(
    client: AsyncClient,
) -> None:
    headers = await _auth_headers(client)
    file_bytes = _build_salary_workbook(
        [
            {
                "employee_name": "Scientific IDs",
                "work_permit_no": "6.17311E+06",
                "personal_no": "1.0010127744762E+13",
                "bank_name_routing_no": "Bank Z",
                "bank_account_no": "ACC999",
                "days_absent": 0,
                "fixed_portion": "1000.00",
                "variable_portion": "0.00",
                "total_payment": "1000.00",
            }
        ]
    )
    import_response = await client.post(
        "/api/v1/employee-salaries/import",
        headers=headers,
        files={
            "file": (
                "salary_import_scientific.xlsx",
                file_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert import_response.status_code == 200
    assert import_response.json()["created"] == 1
    assert import_response.json()["updated"] == 0


@pytest.mark.asyncio
async def test_import_employee_salaries_auto_trims_over_length_ids(client: AsyncClient) -> None:
    headers = await _auth_headers(client)
    file_bytes = _build_salary_workbook(
        [
            {
                "employee_name": "Trimmed IDs",
                "work_permit_no": "9912345678",
                "personal_no": "8812345678901234",
                "bank_name_routing_no": "Bank T",
                "bank_account_no": "ACC111",
                "days_absent": 0,
                "fixed_portion": "900.00",
                "variable_portion": "10.00",
                "total_payment": "910.00",
            }
        ]
    )
    import_response = await client.post(
        "/api/v1/employee-salaries/import",
        headers=headers,
        files={
            "file": (
                "salary_import_trimmed.xlsx",
                file_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert import_response.status_code == 200
    assert import_response.json()["created"] == 1
    assert import_response.json()["updated"] == 0

    rows = await client.get("/api/v1/employee-salaries", headers=headers)
    assert rows.status_code == 200
    payload = rows.json()
    assert len(payload) == 1
    assert payload[0]["work_permit_no"] == "12345678"
    assert payload[0]["personal_no"] == "12345678901234"
