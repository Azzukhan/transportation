from __future__ import annotations

import math
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from src.core.exceptions import AppException
from src.models.employee_salary import EmployeeSalary

MONTH_NAMES = ["JAN", "FEB", "MAR", "APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"]
WORK_PERMIT_LENGTH = 8
PERSONAL_NO_LENGTH = 14


@dataclass
class RowStyleTemplate:
    data_row_styles: dict[int, object]  # col_index -> style proxy
    total_row_styles: dict[int, object]


@dataclass
class ParsedEmployeeSalaryRow:
    row_number: int
    employee_name: str
    work_permit_no: str
    personal_no: str
    bank_name_routing_no: str | None
    bank_account_no: str
    days_absent: int | None
    fixed_portion: Decimal
    variable_portion: Decimal
    total_payment: Decimal | None
    on_leave: bool


@dataclass
class ParsedEmployeeSalarySkippedRow:
    row_number: int
    reason: str


@dataclass
class ParsedEmployeeSalarySheet:
    rows: list[ParsedEmployeeSalaryRow]
    skipped: list[ParsedEmployeeSalarySkippedRow]


class EmployeeSalaryExcelService:
    TEMPLATE_PATH = "src/assets/templates/employee_salary_template.xlsx"

    @classmethod
    def generate_salary_sheet(
        cls,
        employees: list[EmployeeSalary],
        month: int,
        year: int,
    ) -> bytes:
        wb = load_workbook(cls.TEMPLATE_PATH)
        ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active

        month_name = MONTH_NAMES[month - 1]
        ws.title = f"{month_name} Sikar Cargo Salary Part 1"

        # Header
        header_text = f"PAYROLL FOR THE MONTH OF {month_name} - {year}"
        ws["D3"].value = header_text

        visible = [e for e in employees if not e.on_leave]

        # Template rows in your sheet
        DATA_TEMPLATE_ROW = 9
        TOTAL_TEMPLATE_ROW = 24

        tmpl = cls._extract_style_template(ws, DATA_TEMPLATE_ROW, TOTAL_TEMPLATE_ROW)

        # Your original template seems to have:
        # - rows <= 8 as header part
        # - rows 9..23 data
        # - row 24 totals
        # - rows >= 25 footer/notes/etc
        #
        # We will:
        # 1) delete existing data rows 9..24
        # 2) insert needed rows for data + total
        #
        # Count how many rows to create
        start_row = 9
        data_count = len(visible)
        total_row = start_row + data_count

        # Delete old block (9..24 inclusive) safely
        ws.delete_rows(idx=9, amount=(24 - 9 + 1))

        # Insert new block for data + total
        ws.insert_rows(idx=9, amount=(data_count + 1))  # +1 for total row

        # Fill data rows
        for i, emp in enumerate(visible, start=1):
            r = start_row + (i - 1)
            cls._write_salary_row(ws, r, i, emp, tmpl)

        # Fill total row
        cls._write_total_row(ws, total_row, visible, tmpl)

        # (Optional) Update worksheet dimension/print area if you rely on it
        # ws.calculate_dimension() is automatic in openpyxl when saved.
        # If your template has a defined print area, you can adjust it similarly.

        out = BytesIO()
        wb.save(out)
        return out.getvalue()

    @staticmethod
    def _extract_style_template(ws: Worksheet, data_row: int, total_row: int) -> RowStyleTemplate:
        # Copy styles from template rows (column 1..26 => A..Z)
        data_styles = {}
        total_styles = {}
        for col in range(1, 27):
            data_styles[col] = ws.cell(row=data_row, column=col)._style
            total_styles[col] = ws.cell(row=total_row, column=col)._style
        return RowStyleTemplate(data_row_styles=data_styles, total_row_styles=total_styles)

    @staticmethod
    def _apply_style(cell: Cell, style_obj: object) -> None:
        # openpyxl stores internal style objects; assign directly
        cell._style = style_obj

    @classmethod
    def _write_salary_row(
        cls, ws: Worksheet, row: int, serial: int, emp: EmployeeSalary, tmpl: RowStyleTemplate
    ) -> None:
        # Apply row styles first (A..Z)
        for col in range(1, 27):
            c = ws.cell(row=row, column=col)
            cls._apply_style(c, tmpl.data_row_styles.get(col, c._style))

        # Values (only the columns you use)
        ws.cell(row=row, column=1, value=serial)  # A
        ws.cell(row=row, column=2, value=emp.employee_name)  # B
        ws.cell(row=row, column=3, value=emp.work_permit_no)  # C
        ws.cell(row=row, column=4, value=emp.personal_no)  # D
        ws.cell(row=row, column=5, value=(emp.bank_name_routing_no or ""))  # E
        ws.cell(row=row, column=6, value=emp.bank_account_no)  # F

        # Days absent can be blank
        if emp.days_absent is not None:
            ws.cell(row=row, column=7, value=int(emp.days_absent))  # G

        # Decimals: write as float or Decimal (openpyxl supports Decimal but float is safer for Excel display)
        ws.cell(row=row, column=8, value=float(emp.fixed_portion))  # H
        ws.cell(row=row, column=9, value=float(emp.variable_portion))  # I
        ws.cell(row=row, column=10, value=float(emp.total_payment))  # J

    @classmethod
    def _write_total_row(
        cls, ws: Worksheet, row: int, emps: list[EmployeeSalary], tmpl: RowStyleTemplate
    ) -> None:
        for col in range(1, 27):
            c = ws.cell(row=row, column=col)
            cls._apply_style(c, tmpl.total_row_styles.get(col, c._style))

        fixed_total = sum((e.fixed_portion for e in emps), Decimal("0.00"))
        variable_total = sum((e.variable_portion for e in emps), Decimal("0.00"))
        payment_total = sum((e.total_payment for e in emps), Decimal("0.00"))

        ws.cell(row=row, column=8, value=float(fixed_total))  # H
        ws.cell(row=row, column=9, value=float(variable_total))  # I
        ws.cell(row=row, column=10, value=float(payment_total))  # J

    @classmethod
    def parse_salary_sheet(cls, file_bytes: bytes) -> ParsedEmployeeSalarySheet:
        if not file_bytes:
            raise AppException("Empty file upload is not allowed", status_code=400)
        try:
            wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
        except Exception as exc:  # pragma: no cover - openpyxl exception types vary
            raise AppException("Invalid Excel file", status_code=400) from exc

        ws = wb.active
        parsed_rows: list[ParsedEmployeeSalaryRow] = []
        skipped_rows: list[ParsedEmployeeSalarySkippedRow] = []
        # Template data rows start from row 9, columns B..J hold salary fields.
        for row in range(9, ws.max_row + 1):
            employee_name = cls._normalize_text(ws.cell(row=row, column=2).value)
            work_permit_no = cls._normalize_digit_string(ws.cell(row=row, column=3).value)
            personal_no = cls._normalize_digit_string(ws.cell(row=row, column=4).value)
            if work_permit_no:
                work_permit_no = cls._fit_digit_length(work_permit_no, WORK_PERMIT_LENGTH)
            if personal_no:
                personal_no = cls._fit_digit_length(personal_no, PERSONAL_NO_LENGTH)
            bank_name_routing_no = cls._normalize_text(ws.cell(row=row, column=5).value)
            bank_account_no = cls._normalize_text(ws.cell(row=row, column=6).value) or ""
            try:
                days_absent = cls._to_int_or_none(ws.cell(row=row, column=7).value)
                fixed_portion = cls._to_decimal(ws.cell(row=row, column=8).value)
                variable_portion = cls._to_decimal(ws.cell(row=row, column=9).value)
                total_payment = cls._to_decimal_or_none(ws.cell(row=row, column=10).value)
            except ArithmeticError:
                skipped_rows.append(
                    ParsedEmployeeSalarySkippedRow(row_number=row, reason="invalid numeric values")
                )
                continue

            if cls._is_empty_row(
                employee_name,
                work_permit_no,
                personal_no,
                bank_name_routing_no,
                bank_account_no,
                days_absent,
                fixed_portion,
                variable_portion,
                total_payment,
            ):
                continue
            if employee_name and employee_name.upper() == "TOTAL":
                continue
            if not employee_name:
                skipped_rows.append(
                    ParsedEmployeeSalarySkippedRow(
                        row_number=row, reason="employee_name is required"
                    )
                )
                continue
            if not work_permit_no:
                skipped_rows.append(
                    ParsedEmployeeSalarySkippedRow(
                        row_number=row, reason="work_permit_no is required"
                    )
                )
                continue
            if not personal_no:
                skipped_rows.append(
                    ParsedEmployeeSalarySkippedRow(row_number=row, reason="personal_no is required")
                )
                continue
            if not bank_account_no:
                skipped_rows.append(
                    ParsedEmployeeSalarySkippedRow(
                        row_number=row, reason="bank_account_no is required"
                    )
                )
                continue
            if len(work_permit_no) != WORK_PERMIT_LENGTH or len(personal_no) != PERSONAL_NO_LENGTH:
                work_msg = (
                    f"work_permit_no has {len(work_permit_no)} digits (expected {WORK_PERMIT_LENGTH})"
                    if len(work_permit_no) != WORK_PERMIT_LENGTH
                    else None
                )
                personal_msg = (
                    f"personal_no has {len(personal_no)} digits (expected {PERSONAL_NO_LENGTH})"
                    if len(personal_no) != PERSONAL_NO_LENGTH
                    else None
                )
                mismatch_reason = " and ".join(
                    part for part in [work_msg, personal_msg] if part is not None
                )
                skipped_rows.append(
                    ParsedEmployeeSalarySkippedRow(
                        row_number=row,
                        reason=mismatch_reason,
                    )
                )
                continue
            if days_absent is not None and (days_absent < 0 or days_absent > 31):
                skipped_rows.append(
                    ParsedEmployeeSalarySkippedRow(
                        row_number=row, reason="days_absent must be 0..31"
                    )
                )
                continue

            parsed_rows.append(
                ParsedEmployeeSalaryRow(
                    row_number=row,
                    employee_name=employee_name,
                    work_permit_no=work_permit_no,
                    personal_no=personal_no,
                    bank_name_routing_no=bank_name_routing_no,
                    bank_account_no=bank_account_no,
                    days_absent=days_absent,
                    fixed_portion=fixed_portion,
                    variable_portion=variable_portion,
                    total_payment=total_payment,
                    on_leave=False,
                )
            )
        if not parsed_rows:
            raise AppException("No valid employee salary rows found in Excel file", status_code=400)
        return ParsedEmployeeSalarySheet(rows=parsed_rows, skipped=skipped_rows)

    @staticmethod
    def _normalize_text(value: Any) -> str | None:
        if value is None:
            return None
        if isinstance(value, str):
            cleaned = value.strip()
            return cleaned or None
        cleaned = str(value).strip()
        return cleaned or None

    @classmethod
    def _normalize_digit_string(cls, value: Any) -> str | None:
        if value is None:
            return None

        if isinstance(value, int):
            return str(abs(value))

        if isinstance(value, float):
            if not math.isfinite(value):
                return None
            return str(abs(int(round(value))))

        if isinstance(value, Decimal):
            try:
                return str(abs(int(value.to_integral_value())))
            except (OverflowError, InvalidOperation):
                return None

        text = cls._normalize_text(value)
        if text is None:
            return None

        compact = text.replace(",", "").replace(" ", "")
        if compact.endswith(".0"):
            compact = compact[:-2]
        if compact.isdigit():
            return compact

        # Scientific notation or decimal-looking text from Excel exports.
        try:
            numeric = Decimal(compact)
            return str(abs(int(numeric.to_integral_value())))
        except (InvalidOperation, OverflowError):
            pass

        digits = "".join(ch for ch in compact if ch.isdigit())
        return digits or None

    @staticmethod
    def _fit_digit_length(value: str, expected_length: int) -> str:
        if len(value) > expected_length:
            value = value[-expected_length:]
        return value.zfill(expected_length)

    @staticmethod
    def _to_int_or_none(value: Any) -> int | None:
        if value is None or value == "":
            return None
        return int(Decimal(str(value)))

    @staticmethod
    def _to_decimal(value: Any) -> Decimal:
        if value is None or value == "":
            return Decimal("0.00")
        return Decimal(str(value)).quantize(Decimal("0.01"))

    @staticmethod
    def _to_decimal_or_none(value: Any) -> Decimal | None:
        if value is None or value == "":
            return None
        return Decimal(str(value)).quantize(Decimal("0.01"))

    @staticmethod
    def _is_empty_row(
        employee_name: str | None,
        work_permit_no: str | None,
        personal_no: str | None,
        bank_name_routing_no: str | None,
        bank_account_no: str,
        days_absent: int | None,
        fixed_portion: Decimal,
        variable_portion: Decimal,
        total_payment: Decimal | None,
    ) -> bool:
        return (
            employee_name is None
            and work_permit_no is None
            and personal_no is None
            and bank_name_routing_no is None
            and bank_account_no == ""
            and days_absent is None
            and fixed_portion == Decimal("0.00")
            and variable_portion == Decimal("0.00")
            and total_payment is None
        )
